"""
Provider Alerts Evaluation Script

This script reads a consolidated provider operations CSV and
evaluates technical failure and low-volume alerts using
hardcoded thresholds.

Output is printed to console for exploratory and demo purposes.
"""

import os
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


# =========================
# Configuration
# =========================

load_dotenv()

INPUT_CSV = os.getenv("SUMMARY_OUTPUT_CSV_URL")
OUTPUT_EXCEL = os.getenv("PROVIDER_ALERTS_EXCEL")

# =========================
# Alert evaluation logic
# =========================

def evaluate_failure_alert(total_operations: int, failure_rate: float
) -> str:
    """
    Evaluate technical failure alert status.

    Returns:
        str: Alert level
    """
    if total_operations < 500:
        return "CAN'T EVALUATE"

    if failure_rate < 0.10:
        return "NORMAL"
    if failure_rate <= 0.25:
        return "CONCERN"

    return "SEVERE"


def evaluate_volume_alert(total_operations: int) -> str:
    """
    Evaluate low-volume alert status.

    Returns:
        str: Alert level
    """
    if total_operations < 100:
        return "URGENT"
    if total_operations < 500:
        return "SEVERE"
    if total_operations < 2000:
        return "CONCERN"

    return "NORMAL"

# =========================
# Color mapping
# =========================

ALERT_COLORS = {
    "NORMAL": PatternFill("solid", fgColor="C6EFCE"),       # Green
    "CONCERN": PatternFill("solid", fgColor="FFEE00"),     # Yellow
    "SEVERE": PatternFill("solid", fgColor="FF2900"),       # Red
    "URGENT": PatternFill("solid", fgColor="751400"),      # Dark Red
    "CAN'T EVALUATE": PatternFill("solid", fgColor="D9D9D9")  # Grey
}

# =========================
# Main processing
# =========================

def main() -> None:
    """
    Main execution flow for alert evaluation.
    """
    df = pd.read_csv(INPUT_CSV)

    # Evaluate alerts
    df["failure_alert"] = df.apply(
        lambda r: evaluate_failure_alert(
            int(r["total_operations"]),
            float(r["failure_rate"])
        ),
        axis=1
    )

    df["volume_alert"] = df["total_operations"].apply(
        evaluate_volume_alert
    )

    # Export to Excel
    df.to_excel(OUTPUT_EXCEL, index=False)

    # Apply colors
    workbook = load_workbook(OUTPUT_EXCEL)
    sheet = workbook.active

    header = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}

    for row in range(2, sheet.max_row + 1):
        for col_name in ["failure_alert", "volume_alert"]:
            col_idx = header[col_name]
            cell = sheet.cell(row=row, column=col_idx)
            fill = ALERT_COLORS.get(cell.value)
            if fill:
                cell.fill = fill

    workbook.save(OUTPUT_EXCEL)

    print(f"Reporte de alertas generado: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()
