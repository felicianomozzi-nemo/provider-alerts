"""
Provider Alerts Evaluation Script

This script reads a consolidated provider operations CSV and
evaluates technical failure and low-volume alerts using
hardcoded thresholds.

Output is printed to console for exploratory and demo purposes.
"""

import os
import sys
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


# =========================
# Configuration
# =========================

load_dotenv()
SUMMARY_DIR = os.getenv("SUMMARY_DIR")
HISTORIC_SUMMARY_DIR = os.getenv("HISTORIC_SUMMARY_DIR")
SUMMARY_ENDING = os.getenv("SUMMARY_ENDING")

INPUT_FILES = {
    "provider_name": os.getenv("PROVIDERS_SUMMARY_CSV"),
    "hotel_name": os.getenv("HOTELS_SUMMARY_CSV"),
    "client_name": os.getenv("CLIENTS_SUMMARY_CSV"),
    "destination_name": os.getenv("DESTINATIONS_SUMMARY_CSV"),
}

HISTORIC_FILES = {
    "provider_name": os.getenv("HISTORIC_PROVIDERS_SUMMARY_CSV"),
    "hotel_name": os.getenv("HISTORIC_HOTELS_SUMMARY_CSV"),
    "client_name": os.getenv("HISTORIC_CLIENTS_SUMMARY_CSV"),
    "destination_name": os.getenv("HISTORIC_DESTINATIONS_SUMMARY_CSV"),
}

OUTPUT_DIR = os.getenv("OUTPUT_DIR")
OUTPUT_ENDING = os.getenv("OUTPUT_ENDING")

# =========================
# Alert evaluation logic
# =========================
def classify_failure_alert(deviation: float) -> str:
    """
    Returns the alert for the failure rate
    """
    if pd.isna(deviation):
        return "CAN'T EVALUATE"
    if deviation <= 0.25:
        return "NORMAL"
    elif deviation <= 0.50:
        return "CONCERN"
    elif deviation <= 0.75:
        return "SEVERE"
    else:
        return "URGENT"

def classify_volume_alert(deviation: float) -> str:
    """
    Returns the alert for the volume
    """
    if pd.isna(deviation):
        return "CAN'T EVALUATE"
    if deviation >= -0.25:
        return "NORMAL"
    elif deviation >= -0.50:
        return "CONCERN"
    elif deviation >= -0.75:
        return "SEVERE"
    else:
        return "URGENT"
def evaluate_failure_alert(df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluate technical failure alert status.

    Returns:
        DataFrame: DataFrame with the new columns
    """
    df["failure_deviation"] = (
        (df["failure_rate"] - df["failure_rate_historic"]) / df["failure_rate_historic"])
    df["failure_alert"] = df["failure_deviation"].apply(classify_failure_alert)

    return df


def evaluate_volume_alert(df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluate low-volume alert status.

    Returns:
        DataFrame: DataFrame with the new columns
    """
    df["volume_deviation"] = (
        (df["daily_avg_operations"] - df["daily_avg_operations_historic"]) / df["daily_avg_operations_historic"]
    )

    df["volume_difference_absolute"] = (
        df["daily_avg_operations"] - df["daily_avg_operations_historic"]
    )

    df["volume_alert"] = df["volume_deviation"].apply(classify_volume_alert)

    return df


# =========================
# Color mapping
# =========================

ALERT_COLORS = {
    "NORMAL": PatternFill("solid", fgColor="C6EFCE"),       # Green
    "CONCERN": PatternFill("solid", fgColor="FFEE00"),     # Yellow
    "SEVERE": PatternFill("solid", fgColor="FF2900"),       # Red
    "URGENT": PatternFill("solid", fgColor="751400"),      # Dark Red
    "CAN'T EVALUATE": PatternFill("solid", fgColor="D9D9D9")  # Grey
}

# =========================
# Main processing
# =========================

def main() -> None:
    """
    Main execution flow for alert evaluation.
    """
    if len(sys.argv) > 1:
        group_by_column = sys.argv[1]
    else:
        group_by_column = "provider_name" #default
    if len(sys.argv) > 2:
        time_range = sys.argv[2]
    else:
        time_range = "now-7d" #default

    if group_by_column not in HISTORIC_FILES:
        raise ValueError(f"Invalid group_by_column: {group_by_column}")
    input_csv = SUMMARY_DIR + INPUT_FILES[group_by_column] + time_range + SUMMARY_ENDING
    historic_csv = HISTORIC_SUMMARY_DIR + HISTORIC_FILES[group_by_column] + SUMMARY_ENDING

    df_current = pd.read_csv(input_csv)
    df_historic = pd.read_csv(historic_csv)

    df = df_current.merge(
        df_historic,
        on=group_by_column,
        how="left",
        suffixes=("", "_historic")
    )

    df = evaluate_failure_alert(df)
    df = evaluate_volume_alert(df)

    df.replace([float("inf"), -float("inf")], pd.NA, inplace=True)

    columns_to_export = [
        group_by_column,
        "failure_rate",
        "failure_rate_historic",
        "failure_deviation",
        "failure_alert",
        "daily_avg_operations",
        "daily_avg_operations_historic",
        "volume_difference_absolute",
        "volume_deviation",
        "volume_alert"
    ]

    df = df[columns_to_export]

    df["failure_rate"] = df["failure_rate"].round(4)
    df["failure_rate_historic"] = df["failure_rate_historic"].round(4)
    df["failure_deviation"] = df["failure_deviation"].round(4)
    df["daily_avg_operations"] = df["daily_avg_operations"].round(0)
    df["daily_avg_operations_historic"] = df["daily_avg_operations_historic"].round(0)
    df["volume_difference_absolute"] = df["volume_difference_absolute"].round(0)
    df["volume_deviation"] = df["volume_deviation"].round(4)

    df = df.rename(columns = {
        group_by_column: "Entidad",
        "failure_rate": "Ratio de Fallos",
        "failure_rate_historic": "Ratio de Fallos Histórico",
        "failure_deviation": "Desviación",
        "failure_alert": "Alerta Fallos",
        "daily_avg_operations": "Operaciones Diarias Promedio",
        "daily_avg_operations_historic": "Operaciones Diarias Promedio Históricas",
        "volume_deviation": "Desviación",
        "volume_difference_absolute": "Diferencia Absoluta",
        "volume_alert": "Alerta Volumen"
    })

    # Export to Excel
    output_path = OUTPUT_DIR + group_by_column + "_" + time_range + OUTPUT_ENDING
    df.to_excel(output_path, index=False)

    # Apply colors
    workbook = load_workbook(output_path)
    sheet = workbook.active

    header = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}

    for row in range(2, sheet.max_row + 1):
        for col_name in ["Alerta Fallos", "Alerta Volumen"]:
            col_idx = header[col_name]
            cell = sheet.cell(row=row, column=col_idx)
            fill = ALERT_COLORS.get(cell.value)
            if fill:
                cell.fill = fill

    workbook.save(output_path)

    print(f"Reporte de alertas generado: {output_path}")


if __name__ == "__main__":
    main()
