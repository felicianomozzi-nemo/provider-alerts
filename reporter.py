"""
Reporting Module.
Takes the evaluated alert datasets, formats columns and values, applies styling,
and exports a final multi-sheet Excel report.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

def format_dataframes(alerts_dict: dict, group_by_column: str) -> dict:
    """
    Renames columns for human readability and rounds numerical values.
    
    Args:
        alerts_dict (dict): Dictionary containing raw alert DataFrames.
        group_by_column (str): The main entity column name.

    Returns:
        dict: A dictionary of formatted DataFrames mapped by Sheet Name.
    """
    df_failure = alerts_dict["failure"].copy()
    df_volume = alerts_dict["volume"].copy()
    df_version = alerts_dict["version"].copy()
    df_summary = alerts_dict["summary"].copy()

    # Rename & Round Failure Alerts
    cols_failure = [
        group_by_column, "politics_search", "loi", "species", "mamushka",
        "failure_rate", "failure_rate_historic", "failure_alert"
    ]
    df_failure = df_failure[[c for c in cols_failure if c in df_failure.columns]]
    df_failure = df_failure.rename(columns={
        group_by_column: "Entity Name", "politics_search": "Politics Search",
        "loi": "LOI", "species": "Species", "mamushka": "Mamushka",
        "failure_rate": "Failure Rate %", "failure_rate_historic": "Historic Failure Rate %",
        "failure_alert": "Failure Alert"
    }).round({"Failure Rate %": 2, "Historic Failure Rate %": 2})

    # Rename & Round Volume Alerts
    cols_volume = [
        group_by_column, "politics_search", "loi", "species", "mamushka",
        "daily_avg_operations", "daily_avg_operations_historic", "volume_deviation", "volume_alert"
    ]
    df_volume = df_volume[[c for c in cols_volume if c in df_volume.columns]]
    df_volume = df_volume.rename(columns={
        group_by_column: "Entity Name", "politics_search": "Politics Search",
        "loi": "LOI", "species": "Species", "mamushka": "Mamushka",
        "daily_avg_operations": "Current Daily Avg Ops", 
        "daily_avg_operations_historic": "Historic Daily Avg Ops",
        "volume_deviation": "Volume Deviation %", "volume_alert": "Volume Alert"
    }).round({"Current Daily Avg Ops": 0, "Historic Daily Avg Ops": 0, "Volume Deviation %": 2})

    # Rename Version Alerts
    cols_version = [
        group_by_column, "politics_search", "loi", "species", "mamushka", "total_operations"
    ]
    df_version = df_version[[c for c in cols_version if c in df_version.columns]]
    df_version = df_version.rename(columns={
        group_by_column: "Entity Name", "politics_search": "Politics Search",
        "loi": "LOI", "species": "Species", "mamushka": "Mamushka",
        "total_operations": "Operations in Period"
    })

    return {
        "Summary": df_summary,
        "Failure Alerts": df_failure,
        "Volume Alerts": df_volume,
        "Invalid Entities": df_version
    }

def style_excel_workbook(output_path: str) -> None:
    """Applies openpyxl styles to the generated Excel workbook."""
    workbook = load_workbook(output_path)
    header_font = Font(name="Montserrat", bold=True, size=12)
    cell_font = Font(name="Montserrat", size=10)
    row_border = Border(bottom=Side(style="thin", color="D9D9D9"))

    header_colors = {
        "Entity Name": PatternFill("solid", fgColor="BDD7EE"),
        "Politics Search": PatternFill("solid", fgColor="D9E1F2"),
        "LOI": PatternFill("solid", fgColor="D9E1B2"),
        "Species": PatternFill("solid", fgColor="E7E6E6"),
        "Mamushka": PatternFill("solid", fgColor="D9F1D5"),
        "Failure Rate %": PatternFill("solid", fgColor="F4B084"),
        "Historic Failure Rate %": PatternFill("solid", fgColor="F8CBFA"),
        "Failure Alert": PatternFill("solid", fgColor="FFF2CC"),
        "Current Daily Avg Ops": PatternFill("solid", fgColor="DDEBF7"),
        "Historic Daily Avg Ops": PatternFill("solid", fgColor="DDECD1"),
        "Volume Deviation %": PatternFill("solid", fgColor="C6E0B4"),
        "Volume Alert": PatternFill("solid", fgColor="FFF2CC"),
        "Operations in Period": PatternFill("solid", fgColor="DEDAEF"),
        "Metric": PatternFill("solid", fgColor="E7E6E6"),
        "Value": PatternFill("solid", fgColor="BDD7EE"),
    }

    row_soft_colors = {
        "Entity Name": PatternFill("solid", fgColor="E8F1FB"),
        "Politics Search": PatternFill("solid", fgColor="EEF2FB"),
        "LOI": PatternFill("solid", fgColor="F3F6E3"),
        "Species": PatternFill("solid", fgColor="F4F4F4"),
        "Mamushka": PatternFill("solid", fgColor="EEF9EC"),
        "Operations in Period": PatternFill("solid", fgColor="F1EFFB"),
        "Metric": PatternFill("solid", fgColor="F2F2F2"),
        "Value": PatternFill("solid", fgColor="E8F1FB"),
    }

    alert_colors = {
        "CONCERN": PatternFill("solid", fgColor="FFF2CC"),
        "SEVERE": PatternFill("solid", fgColor="F8CBAD"),
        "URGENT": PatternFill("solid", fgColor="F4B084")
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]

        # Header Styles
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = row_border
            if header in header_colors:
                cell.fill = header_colors[header]

        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "B2"

        alert_column = next((idx + 1 for idx, name in enumerate(headers) if "Alert" in str(name)), None)

        # Row Styles
        for row in range(2, sheet.max_row + 1):
            alert_fill = None
            if alert_column:
                alert_value = sheet.cell(row=row, column=alert_column).value
                alert_fill = alert_colors.get(alert_value)

            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                header_name = headers[col-1]

                cell.font = cell_font
                cell.border = row_border

                if header_name != "Entity Name":
                    cell.alignment = Alignment(horizontal="center")

                if alert_fill:
                    cell.fill = alert_fill
                elif sheet_name in ["Summary", "Invalid Entities"]:
                    soft_fill = row_soft_colors.get(header_name)
                    if soft_fill:
                        cell.fill = soft_fill

        # Auto-adjust column width
        for col in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = max((len(str(sheet.cell(row=row, column=col).value)) for row in range(1, sheet.max_row + 1) if sheet.cell(row=row, column=col).value), default=0)
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40) + 8

    workbook.save(output_path)

def run_reporting(alerts_dict: dict, group_by_column: str, time_range: str) -> None:
    """
    Main execution flow for generating the Excel report.
    
    Args:
        alerts_dict (dict): Evaluated alerts from the alerter module.
        group_by_column (str): The column used to aggregate the data.
        time_range (str): Processed time range.
    """
    formatted_dfs = format_dataframes(alerts_dict, group_by_column)
    output_path = f"{config.OUTPUT_DIR}{group_by_column}_{time_range}{config.OUTPUT_ENDING}"
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in formatted_dfs.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    style_excel_workbook(output_path)
    print(f"Excel report visually formatted and saved to: {output_path}")
